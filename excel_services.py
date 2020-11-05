from openpyxl import load_workbook

from config import Config


def get_formated_str(id, record):
    return f"insert into FakeSkill(containerId, orderPrice, shortDesc, description, number, createdAt) VALUES ({id}, null, {record}, null, now());"


def get_data_from_toloka():
    workbook = load_workbook(Config.SOURCE_FILE_NAME + '.xlsx')
    cur_id = None
    new_id = False
    data = ''
    for sheet in workbook:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.column_letter == 'F':
                    if cell.value is not None:
                        if type(cell.value) is int:
                            cur_id = str(cell.value)
                            new_id = True
                        elif type(cell.value) is str:
                            if cell.value.isdigit():
                                cur_id = cell.value
                                new_id = True
                            elif cur_id is not None:
                                data += get_formated_str(cur_id, cell.value) + '\n'
                    elif new_id:
                        new_id = False
                    else:
                        cur_id = None
    return data[:-1]
